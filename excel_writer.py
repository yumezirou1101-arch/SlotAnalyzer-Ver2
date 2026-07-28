# excel_writer.py
"""openpyxlを使ってExcelファイルへデータを保存する処理。"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_HEADERS = [
    "日付",
    "店舗",
    "機種名",
    "台番号",
    "ゲーム数",
    "BB",
    "RB",
    "合成確率",
    "差枚",
]

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)


def style_sheet(sheet) -> None:
    """ヘッダー、フィルター、列幅を設定する。"""
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        sheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            35,
        )


def create_all_data_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    """全台データシートを作成する。"""
    sheet = workbook.active
    sheet.title = "全台データ"

    sheet.append(DATA_HEADERS)

    for row in rows:
        sheet.append(
            [row.get(header) for header in DATA_HEADERS]
        )

    for cell in sheet["A"][1:]:
        if hasattr(cell.value, "strftime"):
            cell.number_format = "yyyy/mm/dd"

    style_sheet(sheet)


def create_machine_summary_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    """機種別集計シートを作成する。"""
    sheet = workbook.create_sheet("機種別集計")

    headers = [
        "日付",
        "店舗",
        "機種名",
        "台数",
        "平均ゲーム数",
        "差枚合計",
        "平均差枚",
    ]

    sheet.append(headers)

    groups: dict[
        tuple[Any, str, str],
        list[dict[str, Any]],
    ] = defaultdict(list)

    for row in rows:
        key = (
            row["日付"],
            row["店舗"],
            row["機種名"],
        )
        groups[key].append(row)

    for key, items in sorted(
        groups.items(),
        key=lambda item: item[0][2],
    ):
        business_date, store_name, machine_name = key

        games = [
            item["ゲーム数"]
            for item in items
            if isinstance(item.get("ゲーム数"), (int, float))
        ]

        medals = [
            item["差枚"]
            for item in items
            if isinstance(item.get("差枚"), (int, float))
        ]

        average_games = None
        total_medals = None
        average_medals = None

        if games:
            average_games = round(
                sum(games) / len(games),
                1,
            )

        if medals:
            total_medals = sum(medals)
            average_medals = round(
                sum(medals) / len(medals),
                1,
            )

        sheet.append(
            [
                business_date,
                store_name,
                machine_name,
                len(items),
                average_games,
                total_medals,
                average_medals,
            ]
        )

    style_sheet(sheet)


def create_ranking_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    """差枚ランキングシートを作成する。"""
    sheet = workbook.create_sheet("差枚ランキング")

    headers = ["順位"] + DATA_HEADERS
    sheet.append(headers)

    ranked_rows = sorted(
        rows,
        key=lambda row: (
            row.get("差枚")
            if isinstance(row.get("差枚"), (int, float))
            else float("-inf")
        ),
        reverse=True,
    )

    for rank, row in enumerate(ranked_rows, start=1):
        sheet.append(
            [rank] + [row.get(header) for header in DATA_HEADERS]
        )

    style_sheet(sheet)


def save_workbook(
    rows: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    """取得データと集計データをExcelファイルへ保存する。"""
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    create_all_data_sheet(workbook, rows)
    create_machine_summary_sheet(workbook, rows)
    create_ranking_sheet(workbook, rows)

    workbook.save(output_path)

    return output_path