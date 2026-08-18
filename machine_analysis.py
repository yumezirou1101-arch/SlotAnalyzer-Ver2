import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ==========================================
# CSV読み込み
# ==========================================

with open(
    "slot_data.csv",
    "r",
    encoding="utf-8-sig"
) as f:

    rows = list(csv.DictReader(f))


# ==========================================
# 数値変換
# ==========================================

def to_int(value):
    try:
        return int(value)
    except:
        return 0


# ==========================================
# 機種別集計
# ==========================================

machines = defaultdict(list)

for row in rows:
    machines[row["機種名"]].append(row)


results = []


for machine, data in machines.items():

    count = len(data)

    g_values = [
        to_int(x["G数"])
        for x in data
    ]

    diff_values = [
        to_int(x["差枚"])
        for x in data
    ]

    bb_values = [
        to_int(x["BB"])
        for x in data
    ]

    rb_values = [
        to_int(x["RB"])
        for x in data
    ]

    # ------------------------------
    # 基本統計
    # ------------------------------

    avg_g = sum(g_values) / count

    avg_diff = sum(diff_values) / count

    avg_bb = sum(bb_values) / count

    avg_rb = sum(rb_values) / count

    # ------------------------------
    # 勝敗
    # ------------------------------

    plus_count = sum(
        x > 0 for x in diff_values
    )

    minus_count = sum(
        x < 0 for x in diff_values
    )

    zero_count = sum(
        x == 0 for x in diff_values
    )

    win_rate = (
        plus_count / count * 100
    )

    # ------------------------------
    # 差枚分布
    # ------------------------------

    plus_1000 = sum(
        x >= 1000
        for x in diff_values
    )

    plus_2000 = sum(
        x >= 2000
        for x in diff_values
    )

    plus_3000 = sum(
        x >= 3000
        for x in diff_values
    )

    minus_1000 = sum(
        x <= -1000
        for x in diff_values
    )

    minus_3000 = sum(
        x <= -3000
        for x in diff_values
    )

    results.append([
        machine,
        count,
        round(avg_g, 1),
        round(avg_diff, 1),
        plus_count,
        minus_count,
        zero_count,
        round(win_rate, 1),
        max(diff_values),
        min(diff_values),
        plus_1000,
        plus_2000,
        plus_3000,
        minus_1000,
        minus_3000,
        round(avg_bb, 1),
        round(avg_rb, 1),
    ])


# ==========================================
# 平均差枚の降順
# ==========================================

results.sort(
    key=lambda x: x[3],
    reverse=True
)


# ==========================================
# Excel作成
# ==========================================

wb = Workbook()

ws = wb.active
ws.title = "機種分析"


headers = [
    "機種名",
    "台数",
    "平均G数",
    "平均差枚",
    "プラス台数",
    "マイナス台数",
    "±0台数",
    "勝率%",
    "最大差枚",
    "最小差枚",
    "+1000枚以上",
    "+2000枚以上",
    "+3000枚以上",
    "-1000枚以下",
    "-3000枚以下",
    "平均BB",
    "平均RB",
]


ws.append(headers)


for row in results:
    ws.append(row)


# ==========================================
# 書式
# ==========================================

for cell in ws[1]:

    cell.font = Font(
        bold=True
    )

    cell.alignment = Alignment(
        horizontal="center"
    )


# フィルター
ws.auto_filter.ref = ws.dimensions

# ウィンドウ固定
ws.freeze_panes = "A2"


# 列幅
widths = {
    "A": 24,
    "B": 8,
    "C": 12,
    "D": 12,
    "E": 12,
    "F": 12,
    "G": 10,
    "H": 10,
    "I": 12,
    "J": 12,
    "K": 13,
    "L": 13,
    "M": 13,
    "N": 13,
    "O": 13,
    "P": 10,
    "Q": 10,
}


for col, width in widths.items():
    ws.column_dimensions[col].width = width


# ==========================================
# 台別データも保存
# ==========================================

ws2 = wb.create_sheet("台別データ")

headers2 = [
    "台番号",
    "機種名",
    "G数",
    "BB",
    "RB",
    "ART",
    "合成",
    "差枚",
]

ws2.append(headers2)

for row in rows:

    ws2.append([
        row["台番号"],
        row["機種名"],
        to_int(row["G数"]),
        to_int(row["BB"]),
        to_int(row["RB"]),
        to_int(row["ART"]),
        row["合成"],
        to_int(row["差枚"]),
    ])


ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = "A2"


# ==========================================
# 保存
# ==========================================

filename = "machine_analysis.xlsx"

wb.save(filename)

print("保存完了")
print("機種数:", len(results))
print("台数:", len(rows))
print("ファイル:", filename)