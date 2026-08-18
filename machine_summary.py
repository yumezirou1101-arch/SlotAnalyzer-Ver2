from openpyxl import load_workbook, Workbook
from collections import defaultdict

# 元データを開く
wb = load_workbook("slot_data.xlsx")
ws = wb.active

# 機種ごとの集計
summary = defaultdict(lambda: {
    "台数": 0,
    "G数": 0,
    "差枚": 0,
    "勝ち台": 0
})

# 2行目からデータ
for row in ws.iter_rows(min_row=2, values_only=True):

    台番号 = row[0]
    機種名 = row[1]
    G数 = row[2]
    BB = row[3]
    RB = row[4]
    合成 = row[5]
    差枚 = row[6]

    try:
        G数 = int(G数)
    except:
        G数 = 0

    try:
        差枚 = int(差枚)
    except:
        差枚 = 0

    summary[機種名]["台数"] += 1
    summary[機種名]["G数"] += G数
    summary[機種名]["差枚"] += 差枚

    if 差枚 > 0:
        summary[機種名]["勝ち台"] += 1

# 新しいExcel
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "機種別集計"

new_ws.append([
    "順位",
    "機種名",
    "台数",
    "平均G数",
    "平均差枚",
    "勝率"
])

# 平均差枚順に並び替え
result = []

for machine, data in summary.items():

    avg_g = round(data["G数"] / data["台数"])

    avg_diff = round(data["差枚"] / data["台数"])

    win_rate = round(data["勝ち台"] / data["台数"] * 100, 1)

    result.append([
        machine,
        data["台数"],
        avg_g,
        avg_diff,
        win_rate
    ])

result.sort(key=lambda x: x[3], reverse=True)

rank = 1

for r in result:

    new_ws.append([
        rank,
        r[0],
        r[1],
        r[2],
        r[3],
        f"{r[4]}%"
    ])

    rank += 1

new_wb.save("machine_summary.xlsx")

print("保存完了")
print("機種数:", len(result))
print("ファイル: machine_summary.xlsx")