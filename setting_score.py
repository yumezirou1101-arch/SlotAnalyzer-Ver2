import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ==========================================
# CSV読み込み
# ==========================================

with open(
    "slot_data.csv",
    "r",
    encoding="utf-8-sig"
) as f:
    rows = list(csv.DictReader(f))


# ==========================================
# 数値変換
# ==========================================

def to_int(value):
    try:
        return int(value)
    except:
        return 0


# ==========================================
# 機種ごとに分類
# ==========================================

machines = defaultdict(list)

for row in rows:
    machines[row["機種名"]].append(row)


# ==========================================
# 設定期待度スコア計算
# ==========================================

results = []


for machine, data in machines.items():

    count = len(data)

    # ------------------------------
    # 機種全体の平均
    # ------------------------------

    avg_g = sum(
        to_int(x["G数"])
        for x in data
    ) / count

    avg_diff = sum(
        to_int(x["差枚"])
        for x in data
    ) / count

    # ------------------------------
    # 機種内の勝率
    # ------------------------------

    plus_count = sum(
        to_int(x["差枚"]) > 0
        for x in data
    )

    machine_win_rate = (
        plus_count / count * 100
    )

    # ------------------------------
    # 機種内順位用
    # ------------------------------

    sorted_data = sorted(
        data,
        key=lambda x: to_int(x["差枚"]),
        reverse=True
    )

    rank_map = {}

    for rank, row in enumerate(
        sorted_data,
        start=1
    ):
        rank_map[row["台番号"]] = rank

    # ------------------------------
    # 各台評価
    # ------------------------------

    for row in data:

        g = to_int(row["G数"])
        diff = to_int(row["差枚"])

        # 機種平均との差
        diff_from_avg = diff - avg_diff

        # ------------------------------
        # G数スコア
        # ------------------------------

        if avg_g > 0:
            g_ratio = g / avg_g
        else:
            g_ratio = 0

        if g_ratio >= 1.3:
            g_score = 20
        elif g_ratio >= 1.1:
            g_score = 15
        elif g_ratio >= 0.9:
            g_score = 10
        elif g_ratio >= 0.7:
            g_score = 5
        else:
            g_score = 0

        # ------------------------------
        # 差枚スコア
        # ------------------------------

        if diff >= 3000:
            diff_score = 30
        elif diff >= 2000:
            diff_score = 25
        elif diff >= 1000:
            diff_score = 20
        elif diff > 0:
            diff_score = 10
        elif diff >= -1000:
            diff_score = 3
        else:
            diff_score = 0

        # ------------------------------
        # 機種平均との差スコア
        # ------------------------------

        if diff_from_avg >= 3000:
            relative_score = 20
        elif diff_from_avg >= 2000:
            relative_score = 15
        elif diff_from_avg >= 1000:
            relative_score = 10
        elif diff_from_avg >= 0:
            relative_score = 5
        else:
            relative_score = 0

        # ------------------------------
        # 合計スコア
        # ------------------------------

        score = (
            g_score
            + diff_score
            + relative_score
        )

        # ------------------------------
        # 評価
        # ------------------------------

        if score >= 65:
            evaluation = "S"
        elif score >= 50:
            evaluation = "A"
        elif score >= 35:
            evaluation = "B"
        elif score >= 20:
            evaluation = "C"
        else:
            evaluation = "D"

        results.append([
            row["台番号"],
            machine,
            g,
            diff,
            round(avg_g, 1),
            round(avg_diff, 1),
            round(diff_from_avg, 1),
            round(g_ratio, 2),
            machine_win_rate,
            rank_map[row["台番号"]],
            g_score,
            diff_score,
            relative_score,
            score,
            evaluation
        ])


# ==========================================
# スコア順
# ==========================================

results.sort(
    key=lambda x: x[13],
    reverse=True
)


# ==========================================
# Excel作成
# ==========================================

wb = Workbook()

ws = wb.active
ws.title = "設定期待度"


headers = [
    "台番号",
    "機種名",
    "G数",
    "差枚",
    "機種平均G数",
    "機種平均差枚",
    "機種平均との差",
    "G数比",
    "機種勝率%",
    "機種内差枚順位",
    "G数スコア",
    "差枚スコア",
    "相対スコア",
    "総合スコア",
    "評価",
]

ws.append(headers)


for row in results:
    ws.append(row)


# ==========================================
# 書式
# ==========================================

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal="center"
    )


ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"


# 列幅

widths = {
    "A": 10,
    "B": 24,
    "C": 10,
    "D": 10,
    "E": 14,
    "F": 14,
    "G": 14,
    "H": 10,
    "I": 12,
    "J": 15,
    "K": 12,
    "L": 12,
    "M": 12,
    "N": 12,
    "O": 10,
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width


# ==========================================
# 機種別ランキング
# ==========================================

ws2 = wb.create_sheet("機種別ランキング")

ws2.append([
    "機種名",
    "台番号",
    "G数",
    "差枚",
    "総合スコア",
    "評価"
])


for machine, data in machines.items():

    machine_results = [
        x for x in results
        if x[1] == machine
    ]

    machine_results.sort(
        key=lambda x: x[13],
        reverse=True
    )

    for row in machine_results:

        ws2.append([
            row[1],
            row[0],
            row[2],
            row[3],
            row[13],
            row[14]
        ])


ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = "A2"

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 10


# ==========================================
# 上位台だけのランキング
# ==========================================

ws3 = wb.create_sheet("狙い台ランキング")

ws3.append([
    "順位",
    "台番号",
    "機種名",
    "G数",
    "差枚",
    "総合スコア",
    "評価",
])


for rank, row in enumerate(
    results[:50],
    start=1
):

    ws3.append([
        rank,
        row[0],
        row[1],
        row[2],
        row[3],
        row[13],
        row[14]
    ])


ws3.auto_filter.ref = ws3.dimensions
ws3.freeze_panes = "A2"

for col, width in {
    "A": 8,
    "B": 10,
    "C": 24,
    "D": 10,
    "E": 10,
    "F": 12,
    "G": 10,
}.items():
    ws3.column_dimensions[col].width = width


# ==========================================
# 保存
# ==========================================

filename = "setting_score.xlsx"

wb.save(filename)

print("保存完了")
print("台数:", len(results))
print("機種数:", len(machines))
print("ファイル:", filename)

print()
print("===== 上位20台 =====")

for rank, row in enumerate(
    results[:20],
    start=1
):
    print(
        rank,
        row[0],
        row[1],
        "G:", row[2],
        "差枚:", row[3],
        "Score:", row[13],
        "評価:", row[14]
    )