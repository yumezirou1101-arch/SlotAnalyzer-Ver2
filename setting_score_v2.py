import csv
import math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ==========================================
# CSV読み込み
# ==========================================

with open(
    "slot_data.csv",
    "r",
    encoding="utf-8-sig"
) as f:
    rows = list(csv.DictReader(f))


# ==========================================
# 数値変換
# ==========================================

def to_int(value):
    try:
        return int(value)
    except:
        return 0


# ==========================================
# 機種ごとに分類
# ==========================================

machines = defaultdict(list)

for row in rows:
    machines[row["機種名"]].append(row)


# ==========================================
# 機種別統計
# ==========================================

machine_stats = {}

for machine, data in machines.items():

    count = len(data)

    g_values = [
        to_int(x["G数"])
        for x in data
    ]

    diff_values = [
        to_int(x["差枚"])
        for x in data
    ]

    avg_g = sum(g_values) / count

    avg_diff = sum(diff_values) / count

    # 標準偏差
    variance = sum(
        (x - avg_diff) ** 2
        for x in diff_values
    ) / count

    std_diff = math.sqrt(variance)

    plus_count = sum(
        x > 0
        for x in diff_values
    )

    win_rate = plus_count / count * 100

    machine_stats[machine] = {
        "count": count,
        "avg_g": avg_g,
        "avg_diff": avg_diff,
        "std_diff": std_diff,
        "win_rate": win_rate,
    }


# ==========================================
# 各台を評価
# ==========================================

results = []


for machine, data in machines.items():

    stats = machine_stats[machine]

    avg_g = stats["avg_g"]
    avg_diff = stats["avg_diff"]
    std_diff = stats["std_diff"]

    # --------------------------------------
    # 機種内の差枚順位
    # --------------------------------------

    sorted_data = sorted(
        data,
        key=lambda x: to_int(x["差枚"]),
        reverse=True
    )

    rank_map = {}

    for rank, row in enumerate(
        sorted_data,
        start=1
    ):
        rank_map[row["台番号"]] = rank

    # --------------------------------------
    # 各台
    # --------------------------------------

    for row in data:

        dai = row["台番号"]

        g = to_int(row["G数"])
        diff = to_int(row["差枚"])

        # ==================================
        # ① 稼働信頼度
        # ==================================

        if g >= 7000:
            reliability = 1.00
        elif g >= 6000:
            reliability = 0.95
        elif g >= 5000:
            reliability = 0.90
        elif g >= 4000:
            reliability = 0.82
        elif g >= 3000:
            reliability = 0.72
        elif g >= 2000:
            reliability = 0.60
        elif g >= 1000:
            reliability = 0.45
        else:
            reliability = 0.30

        # ==================================
        # ② 差枚実績スコア
        # ==================================

        # 差枚を標準偏差で標準化
        if std_diff > 0:
            z_score = (
                diff - avg_diff
            ) / std_diff
        else:
            z_score = 0

        # 極端な値を制限
        z_score_limited = max(
            -3,
            min(3, z_score)
        )

        # -3～+3 → 0～100
        performance_score = (
            z_score_limited + 3
        ) / 6 * 100

        # ==================================
        # ③ 稼働を考慮した実績スコア
        # ==================================

        weighted_performance = (
            performance_score * reliability
            + 50 * (1 - reliability)
        )

        # ==================================
        # ④ 機種内順位スコア
        # ==================================

        rank = rank_map[dai]

        count = stats["count"]

        if count > 1:
            rank_score = (
                (count - rank)
                / (count - 1)
                * 100
            )
        else:
            rank_score = 50

        # ==================================
        # ⑤ 機種平均との差
        # ==================================

        diff_from_avg = diff - avg_diff

        if std_diff > 0:
            relative_z = (
                diff_from_avg / std_diff
            )
        else:
            relative_z = 0

        relative_z = max(
            -3,
            min(3, relative_z)
        )

        relative_score = (
            relative_z + 3
        ) / 6 * 100

        # ==================================
        # ⑥ 総合スコア
        # ==================================

        total_score = (
            weighted_performance * 0.50
            + rank_score * 0.25
            + relative_score * 0.25
        )

        total_score = round(
            total_score,
            1
        )

        # ==================================
        # 暫定評価
        # ==================================

        if total_score >= 80:
            evaluation = "最上位"
        elif total_score >= 70:
            evaluation = "上位"
        elif total_score >= 60:
            evaluation = "やや上位"
        elif total_score >= 40:
            evaluation = "中位"
        else:
            evaluation = "下位"

        results.append([
            dai,
            machine,
            g,
            diff,
            round(avg_g, 1),
            round(avg_diff, 1),
            round(diff_from_avg, 1),
            round(std_diff, 1),
            round(z_score, 2),
            round(reliability, 2),
            round(performance_score, 1),
            round(weighted_performance, 1),
            rank,
            round(rank_score, 1),
            round(relative_score, 1),
            total_score,
            evaluation,
        ])


# ==========================================
# 総合スコア順
# ==========================================

results.sort(
    key=lambda x: x[15],
    reverse=True
)


# ==========================================
# Excel作成
# ==========================================

wb = Workbook()

ws = wb.active
ws.title = "設定期待度V2"


headers = [
    "台番号",
    "機種名",
    "G数",
    "差枚",
    "機種平均G数",
    "機種平均差枚",
    "平均との差枚",
    "差枚標準偏差",
    "Zスコア",
    "稼働信頼度",
    "実績スコア",
    "信頼度補正後",
    "機種内順位",
    "順位スコア",
    "相対スコア",
    "総合スコア",
    "暫定評価",
]

ws.append(headers)

for row in results:
    ws.append(row)


# ==========================================
# 書式
# ==========================================

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal="center"
    )

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"


widths = {
    "A": 10,
    "B": 24,
    "C": 10,
    "D": 10,
    "E": 14,
    "F": 14,
    "G": 14,
    "H": 14,
    "I": 10,
    "J": 12,
    "K": 12,
    "L": 14,
    "M": 12,
    "N": 12,
    "O": 12,
    "P": 12,
    "Q": 12,
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width


# ==========================================
# 機種別ランキング
# ==========================================

ws2 = wb.create_sheet("機種別ランキング")

ws2.append([
    "機種名",
    "台番号",
    "G数",
    "差枚",
    "機種平均差枚",
    "機種内順位",
    "総合スコア",
    "暫定評価",
])

for machine in machines:

    machine_results = [
        x for x in results
        if x[1] == machine
    ]

    machine_results.sort(
        key=lambda x: x[15],
        reverse=True
    )

    for row in machine_results:

        ws2.append([
            row[1],
            row[0],
            row[2],
            row[3],
            row[5],
            row[12],
            row[15],
            row[16],
        ])

ws2.auto_filter.ref = ws2.dimensions
ws2.freeze_panes = "A2"


for col, width in {
    "A": 24,
    "B": 10,
    "C": 10,
    "D": 10,
    "E": 14,
    "F": 12,
    "G": 12,
    "H": 12,
}.items():
    ws2.column_dimensions[col].width = width


# ==========================================
# 総合ランキング
# ==========================================

ws3 = wb.create_sheet("狙い台ランキング")

ws3.append([
    "順位",
    "台番号",
    "機種名",
    "G数",
    "差枚",
    "機種平均差枚",
    "総合スコア",
    "暫定評価",
])


for rank, row in enumerate(
    results[:50],
    start=1
):

    ws3.append([
        rank,
        row[0],
        row[1],
        row[2],
        row[3],
        row[5],
        row[15],
        row[16],
    ])


ws3.auto_filter.ref = ws3.dimensions
ws3.freeze_panes = "A2"


for col, width in {
    "A": 8,
    "B": 10,
    "C": 24,
    "D": 10,
    "E": 10,
    "F": 14,
    "G": 12,
    "H": 12,
}.items():
    ws3.column_dimensions[col].width = width


# ==========================================
# 保存
# ==========================================

filename = "setting_score_v2.xlsx"

wb.save(filename)


# ==========================================
# コンソール表示
# ==========================================

print("保存完了")
print("台数:", len(results))
print("機種数:", len(machines))
print("ファイル:", filename)

print()
print("===== V2 上位20台 =====")

for rank, row in enumerate(
    results[:20],
    start=1
):

    print(
        rank,
        row[0],
        row[1],
        "G:", row[2],
        "差枚:", row[3],
        "機種平均:", row[5],
        "信頼度:", row[9],
        "Score:", row[15],
        "評価:", row[16]
    )