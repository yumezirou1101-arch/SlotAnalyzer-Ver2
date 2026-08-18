import csv
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ==========================================
# マイジャグラーV 設定スペック
# ==========================================

SETTINGS = {
    1: {"bb": 273.1, "rb": 409.6},
    2: {"bb": 270.8, "rb": 385.5},
    3: {"bb": 266.4, "rb": 336.1},
    4: {"bb": 254.0, "rb": 290.0},
    5: {"bb": 240.1, "rb": 268.6},
    6: {"bb": 229.1, "rb": 229.1},
}


# ==========================================
# CSV読み込み
# ==========================================

with open(
    "slot_data.csv",
    "r",
    encoding="utf-8-sig"
) as f:
    rows = list(csv.DictReader(f))


# ==========================================
# 数値変換
# ==========================================

def to_int(value):
    try:
        return int(value)
    except:
        return 0


# ==========================================
# ポアソン対数尤度
# ==========================================

def poisson_log_likelihood(k, lam):

    if lam <= 0:
        return -999999

    return (
        k * math.log(lam)
        - lam
        - math.lgamma(k + 1)
    )


# ==========================================
# マイVだけ抽出
# ==========================================

target = [
    row
    for row in rows
    if row["機種名"] == "マイV"
]


print("対象台数:", len(target))


# ==========================================
# 各台を設定推測
# ==========================================

results = []


for row in target:

    dai = row["台番号"]

    g = to_int(row["G数"])
    bb = to_int(row["BB"])
    rb = to_int(row["RB"])
    diff = to_int(row["差枚"])

    if g <= 0:
        continue

    likelihoods = {}

    # --------------------------------------
    # 設定1～6の尤度
    # --------------------------------------

    for setting, spec in SETTINGS.items():

        expected_bb = g / spec["bb"]
        expected_rb = g / spec["rb"]

        ll_bb = poisson_log_likelihood(
            bb,
            expected_bb
        )

        ll_rb = poisson_log_likelihood(
            rb,
            expected_rb
        )

        likelihoods[setting] = (
            ll_bb + ll_rb
        )

    # --------------------------------------
    # 数値安定化
    # --------------------------------------

    max_ll = max(
        likelihoods.values()
    )

    weights = {
        setting: math.exp(
            likelihoods[setting] - max_ll
        )
        for setting in likelihoods
    }

    total_weight = sum(
        weights.values()
    )

    probabilities = {
        setting:
        weights[setting] / total_weight * 100
        for setting in SETTINGS
    }

    # --------------------------------------
    # 期待設定
    # --------------------------------------

    expected_setting = sum(
        setting * probabilities[setting]
        for setting in probabilities
    ) / 100

    # --------------------------------------
    # 設定4～6確率
    # --------------------------------------

    high_probability = (
        probabilities[4]
        + probabilities[5]
        + probabilities[6]
    )

    # --------------------------------------
    # 最有力設定
    # --------------------------------------

    best_setting = max(
        probabilities,
        key=probabilities.get
    )

    best_probability = probabilities[
        best_setting
    ]

    # --------------------------------------
    # 実測確率
    # --------------------------------------

    bb_rate = (
        g / bb
        if bb > 0
        else 0
    )

    rb_rate = (
        g / rb
        if rb > 0
        else 0
    )

    combined_count = bb + rb

    combined_rate = (
        g / combined_count
        if combined_count > 0
        else 0
    )

    results.append({
        "台番号": dai,
        "G数": g,
        "BB": bb,
        "RB": rb,
        "差枚": diff,
        "BB確率": bb_rate,
        "RB確率": rb_rate,
        "合成": combined_rate,
        "設定4-6確率": high_probability,
        "期待設定": expected_setting,
        "最有力設定": best_setting,
        "最有力確率": best_probability,
    })


# ==========================================
# 高設定確率順
# ==========================================

results.sort(
    key=lambda x: x["設定4-6確率"],
    reverse=True
)


# ==========================================
# 順位を付ける
# ==========================================

for i, row in enumerate(
    results,
    start=1
):
    row["推測順位"] = i


# ==========================================
# グループ分析
# ==========================================

groups = [
    ("上位5台", results[:5]),
    ("6～10位", results[5:10]),
    ("11～15位", results[10:15]),
    ("16～20位", results[15:20]),
    ("21～28位", results[20:28]),
]


group_results = []


for name, group in groups:

    count = len(group)

    avg_g = sum(
        x["G数"]
        for x in group
    ) / count

    avg_diff = sum(
        x["差枚"]
        for x in group
    ) / count

    avg_high = sum(
        x["設定4-6確率"]
        for x in group
    ) / count

    avg_expected = sum(
        x["期待設定"]
        for x in group
    ) / count

    plus_count = sum(
        x["差枚"] > 0
        for x in group
    )

    win_rate = (
        plus_count / count * 100
    )

    group_results.append([
        name,
        count,
        round(avg_g, 1),
        round(avg_diff, 1),
        round(avg_high, 2),
        round(avg_expected, 2),
        plus_count,
        round(win_rate, 1),
    ])


# ==========================================
# 全体統計
# ==========================================

total_count = len(results)

overall_avg_g = sum(
    x["G数"]
    for x in results
) / total_count

overall_avg_diff = sum(
    x["差枚"]
    for x in results
) / total_count

overall_avg_high = sum(
    x["設定4-6確率"]
    for x in results
) / total_count

overall_avg_expected = sum(
    x["期待設定"]
    for x in results
) / total_count


# ==========================================
# Excel
# ==========================================

wb = Workbook()


# ------------------------------------------
# 台別検証
# ------------------------------------------

ws = wb.active
ws.title = "台別検証"


headers = [
    "推測順位",
    "台番号",
    "G数",
    "BB",
    "RB",
    "差枚",
    "BB確率",
    "RB確率",
    "合成",
    "設定4-6確率%",
    "期待設定",
    "最有力設定",
    "最有力確率%",
]

ws.append(headers)


for row in results:

    ws.append([
        row["推測順位"],
        row["台番号"],
        row["G数"],
        row["BB"],
        row["RB"],
        row["差枚"],
        round(row["BB確率"], 1),
        round(row["RB確率"], 1),
        round(row["合成"], 1),
        round(row["設定4-6確率"], 2),
        round(row["期待設定"], 2),
        row["最有力設定"],
        round(row["最有力確率"], 2),
    ])


# ------------------------------------------
# グループ検証
# ------------------------------------------

ws2 = wb.create_sheet("順位グループ検証")


ws2.append([
    "グループ",
    "台数",
    "平均G数",
    "平均差枚",
    "平均設定4-6確率%",
    "平均期待設定",
    "プラス台数",
    "勝率%",
])


for row in group_results:
    ws2.append(row)


# ------------------------------------------
# 全体統計
# ------------------------------------------

ws3 = wb.create_sheet("全体統計")


ws3.append([
    "項目",
    "値",
])


ws3.append([
    "対象台数",
    total_count,
])

ws3.append([
    "平均G数",
    round(overall_avg_g, 1),
])

ws3.append([
    "平均差枚",
    round(overall_avg_diff, 1),
])

ws3.append([
    "平均設定4-6確率%",
    round(overall_avg_high, 2),
])

ws3.append([
    "平均期待設定",
    round(overall_avg_expected, 2),
])


# ==========================================
# 書式
# ==========================================

for sheet in [ws, ws2, ws3]:

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    sheet.freeze_panes = "A2"


ws.auto_filter.ref = ws.dimensions
ws2.auto_filter.ref = ws2.dimensions


# 列幅

for col, width in {
    "A": 12,
    "B": 10,
    "C": 10,
    "D": 8,
    "E": 8,
    "F": 10,
    "G": 12,
    "H": 12,
    "I": 12,
    "J": 16,
    "K": 12,
    "L": 14,
    "M": 16,
}.items():
    ws.column_dimensions[col].width = width


for col, width in {
    "A": 14,
    "B": 10,
    "C": 12,
    "D": 12,
    "E": 18,
    "F": 14,
    "G": 12,
    "H": 10,
}.items():
    ws2.column_dimensions[col].width = width


# ==========================================
# 保存
# ==========================================

filename = "validate_myv.xlsx"

wb.save(filename)


# ==========================================
# コンソール表示
# ==========================================

print()
print("保存完了")
print("ファイル:", filename)

print()
print("===== 順位グループ検証 =====")

for row in group_results:

    print(
        row[0],
        "台数:", row[1],
        "平均G:", row[2],
        "平均差枚:", row[3],
        "平均設定4-6:", row[4], "%",
        "平均期待設定:", row[5],
        "勝率:", row[7], "%"
    )

print()
print("===== 全体 =====")

print(
    "平均G数:",
    round(overall_avg_g, 1)
)

print(
    "平均差枚:",
    round(overall_avg_diff, 1)
)

print(
    "平均設定4-6確率:",
    round(overall_avg_high, 2),
    "%"
)

print(
    "平均期待設定:",
    round(overall_avg_expected, 2)
)